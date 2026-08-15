"""Reduce Toronto Bike Share trip-level ridership archives to one row per day.

Reads the yearly `bikeshare-ridership-YYYY.{zip,xlsx}` downloads (2016 onward),
keeps only the trip start date, the bike model, and the user type, and writes a
single tidy CSV of daily counts.

The source files are wildly inconsistent, and the inconsistencies are silent:
column names are spelled four different ways; timestamps are ISO in some years
and slash-separated in others; the slash years switch between D/M/YYYY and
M/D/YYYY *between quarters of the same year*; one month is a zip inside a zip;
2016 ships as an xlsx whose second sheet was transposed by Excel on import; and
a few hundred rows are missing a column, shifting every field left.

Rather than assume a layout, each file's date orientation is decided from its
own contents (see `read_csv_stream`), so a format change in a future release
can't quietly move trips to the wrong day.

Output columns:
    date            YYYY-MM-DD
    trips           total trips starting that day
    trips_member    trips by annual/monthly members
    trips_casual    trips by casual users
    trips_member_classic, trips_member_electric,
    trips_casual_classic, trips_casual_electric
                    the user x model joint -- blank before 2024, because the
                    source records no bike model until then

Only what can't be derived is stored. Classic/electric totals are the joint
summed over users (member_classic + casual_classic, and likewise for electric),
so they aren't columns. The member/casual totals *are* columns despite looking
derivable: the joint is blank before 2024, so summing it would silently drop
the 2016-2023 member/casual split. `trips` likewise stays authoritative rather
than being re-derived, so a future row with an unrecognised user type or model
shows up as a shortfall instead of vanishing from the total.

Caveat on the member/casual split: the City's "Annual Member" label decays out
of use from October 2021 and is absent altogether by September 2023, returning
to normal only when the schema changes in January 2024. That whole run is
withheld -- see UNRELIABLE_USER_SPLIT for the evidence. `trips` is unaffected
throughout; only the attribution of trips to rider type is lost.

Usage -- re-run when the City publishes another year, after dropping the new
`bikeshare-ridership-YYYY.zip` into ~/Downloads:

    python scripts/build_bikeshare_daily.py \
        data/bikeshare_daily.csv \
        ../toronto-bike-counters/public/bikeshare-daily.json

Read the run's closing report, not just its exit status: per-year totals, the
count of days with no data, and any "!! dropped rows" note are how a silent
format change shows up.
"""

import csv
import io
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd

DOWNLOADS = Path.home() / "Downloads"
CHUNK = 500_000

# Bike Share Toronto's fleet codes. ICONIC/FIT are the pedal bikes; the E-FIT
# family and ASTRO are pedal-assist.
ELECTRIC_MODELS = ("EFIT", "E-FIT", "ASTRO")

# Anything that opens like a date, in either orientation. The year is 2 or 4
# digits -- 2017 Q4 alone abbreviates it ('10/01/17 00:00:01').
LOOKS_LIKE_DATE = re.compile(r"^\s*\d{1,4}[-/]\d{1,2}[-/]\d{2,4}")
SLASH_DATE = r"^\s*(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})"

START_KEYS = {"starttime", "tripstarttime"}
USER_KEYS = {"usertype"}
MODEL_KEYS = {"bikemodel"}

# date -> Counter of {'trips', 'classic', 'electric', 'member', 'casual'}
tally = defaultdict(Counter)
model_years = set()   # years whose source recorded a bike model at all
notes = Counter()     # anomalies worth printing at the end


def norm(col):
    """Collapse the header spellings ('Trip  Duration', 'Start_Time',
    'trip_start_time', BOM-prefixed 'Start Time') onto one key."""
    return re.sub(r"[\s_]+", "", col.replace("﻿", "").strip().lower())


def classify_model(s):
    """-> Series of 'electric' / 'classic' / NA."""
    u = s.fillna("").str.upper().str.strip()
    electric = u.str.contains("|".join(ELECTRIC_MODELS), regex=True, na=False)
    classic = u.str.contains("ICONIC|FIT", regex=True, na=False) & ~electric
    out = pd.Series(pd.NA, index=s.index, dtype="object")
    out[electric] = "electric"
    out[classic] = "classic"
    return out


def classify_user(s):
    """-> Series of 'member' / 'casual' / NA.

    'Casual Member' (2018-2023) contains both words, so casual is tested first.
    """
    u = s.fillna("").str.lower()
    casual = u.str.contains("casual", na=False)
    member = u.str.contains("member|annual", na=False) & ~casual
    out = pd.Series(pd.NA, index=s.index, dtype="object")
    out[casual] = "casual"
    out[member] = "member"
    return out


def add(target, days, users, models):
    """Fold one batch of already-resolved days into a tally.

    Both marginals are counted, plus the user x model joint. The joint is what
    lets a reader combine the two filters ('member trips on e-bikes'); marginals
    alone can't answer that, and they cover more years, so both are kept.
    """
    for d, n in days.value_counts().items():
        target[d]["trips"] += int(n)
    for src in (users, models):
        for (d, kind), n in days.groupby([days, src]).size().items():
            target[d][kind] += int(n)

    both_known = users.notna() & models.notna()
    if both_known.any():
        pair = users[both_known] + "_" + models[both_known]
        for (d, kind), n in days[both_known].groupby([days[both_known], pair]).size().items():
            target[d][kind] += int(n)


def iso_days(s):
    """Day strings for ISO timestamps; NA where the value isn't parseable."""
    d = pd.to_datetime(s.str.slice(0, 10), format="%Y-%m-%d", errors="coerce")
    return d.dt.strftime("%Y-%m-%d")


def slash_parts(s):
    """-> (year, first, second) as nullable ints for D/M/Y-or-M/D/Y strings."""
    ex = s.str.extract(SLASH_DATE)
    year = ex[2].astype("Int64")
    year = year.where(year >= 100, year + 2000)   # '17' -> 2017
    return year, ex[0].astype("Int64"), ex[1].astype("Int64")


def build_days(year, month, day):
    """Day strings from separate parts, NA where the combination isn't a date
    (e.g. month 13, or Feb 30)."""
    frame = pd.DataFrame({"year": year, "month": month, "day": day})
    valid = frame.notna().all(axis=1) & frame["month"].between(1, 12)
    out = pd.Series(pd.NaT, index=frame.index, dtype="datetime64[ns]")
    if valid.any():
        out[valid] = pd.to_datetime(frame[valid].astype("int64"), errors="coerce")
    return out.dt.strftime("%Y-%m-%d")


def repair_shifted(chunk, col, prev_col, bad):
    """Some rows are missing an earlier field, so every value sits one column to
    the left. Where the start time doesn't look like a date, read across."""
    if prev_col is None or not bad.any():
        return chunk[col]
    return chunk[col].where(~bad, chunk[prev_col])


def read_csv_stream(fh, label):
    """Stream one CSV and fold its daily counts into the global tally.

    Both date orientations are tallied as we go and the file commits to one at
    EOF, decided by whichever component is proven to exceed 12 somewhere in the
    file. That keeps this to a single pass over multi-hundred-MB files while
    still letting each file pick its own format.
    """
    stream = io.TextIOWrapper(fh, encoding="utf-8", errors="replace", newline="")
    header_line = stream.readline()
    if not header_line:
        return
    names = [c.replace("﻿", "").strip()
             for c in next(csv.reader([header_line.rstrip("\r\n")]))]
    cols = {norm(c): c for c in names}
    index = {c: i for i, c in enumerate(names)}

    start_col = next((cols[k] for k in START_KEYS if k in cols), None)
    if start_col is None:
        print(f"  !! no start-time column in {label}: {names}")
        return
    user_col = next((cols[k] for k in USER_KEYS if k in cols), None)
    model_col = next((cols[k] for k in MODEL_KEYS if k in cols), None)

    def before(col):
        i = index[col] - 1 if col else 0
        return names[i] if col and i >= 0 else None

    prev = {c: before(c) for c in (start_col, user_col, model_col) if c}
    usecols = sorted({c for c in list(prev) + list(prev.values()) if c},
                     key=index.get)

    # Tally each orientation separately; pick the winner once the file is read.
    both = {"md": defaultdict(Counter), "dm": defaultdict(Counter)}
    proof = Counter()
    rows = unparsed = 0

    reader = pd.read_csv(stream, header=None, names=names, usecols=usecols,
                         dtype=str, chunksize=CHUNK, on_bad_lines="skip",
                         engine="c")
    for chunk in reader:
        rows += len(chunk)
        raw = chunk[start_col].astype("string").fillna("")
        shifted = ~raw.str.match(LOOKS_LIKE_DATE) & raw.ne("")
        if shifted.any():
            notes[f"field-shifted rows::{label}"] += int(shifted.sum())
            raw = repair_shifted(chunk, start_col, prev[start_col],
                                 shifted).astype("string").fillna("")

        users = (classify_user(repair_shifted(chunk, user_col, prev[user_col],
                                              shifted).astype("string"))
                 if user_col else pd.Series(pd.NA, index=raw.index, dtype="object"))
        models = (classify_model(repair_shifted(chunk, model_col, prev[model_col],
                                                shifted).astype("string"))
                  if model_col else pd.Series(pd.NA, index=raw.index, dtype="object"))

        is_iso = raw.str.len().ge(10) & raw.str.slice(4, 5).eq("-")
        if is_iso.any():
            days = iso_days(raw[is_iso])
            ok = days.notna()
            # Orientation is meaningless here, so both tallies get the same rows.
            for target in both.values():
                add(target, days[ok], users[is_iso][ok], models[is_iso][ok])
            unparsed += int((~ok).sum())

        rest = ~is_iso & raw.ne("")
        if rest.any():
            year, first, second = slash_parts(raw[rest])
            proof["dayfirst"] += int((first > 12).sum())
            proof["monthfirst"] += int((second > 12).sum())
            resolved = {}
            for key, (month, day) in (("md", (first, second)),
                                      ("dm", (second, first))):
                days = build_days(year, month, day)
                ok = days.notna()
                resolved[key] = ok
                add(both[key], days[ok], users[rest][ok], models[rest][ok])
            # Only a row that no orientation can read is genuinely lost; one
            # orientation failing is just that reading being wrong.
            unparsed += int((~resolved["md"] & ~resolved["dm"]).sum())

    if proof["dayfirst"] and proof["monthfirst"]:
        notes[f"!! mixed date orientation::{label}"] = 1
    orientation = "dm" if proof["dayfirst"] >= proof["monthfirst"] and proof["dayfirst"] else "md"

    chosen = both[orientation]
    kept = sum(c["trips"] for c in chosen.values())
    for day, counts in chosen.items():
        tally[day].update(counts)
    if model_col:
        model_years.update(int(d[:4]) for d in chosen)
    if unparsed:
        notes[f"unparsed dates::{label}"] += unparsed
    # Every row read should land on a day. A shortfall means a date shape this
    # parser doesn't know about, which is how a whole quarter goes missing.
    if kept < rows:
        notes[f"!! dropped rows::{label}"] = rows - kept

    shape = ("ISO" if not proof else
             "D/M/Y" if orientation == "dm" else "M/D/Y")
    print(f"  {label}: {kept:,} trips  [{shape}]"
          f"{'' if kept == rows else f'  !! {rows - kept:,} DROPPED'}"
          f"{'' if model_col else '  (no bike model column)'}", flush=True)


def sheet_is_daymonth_swapped(ws, start_idx, sample=50_000):
    """Whether Excel read this sheet's day-first timestamps as month-first.

    The 2016 workbook's Q4 sheet came from a day-first source opened by a
    month-first Excel: every date that survived as a real datetime has its month
    and day transposed (Oct 1 stored as Jan 10), and the ones that couldn't be
    transposed -- day > 12 -- were left behind as text.

    Detected rather than hardcoded: in a swapped sheet the stored *day*
    collapses onto the handful of real months while the stored *month* fans out
    across 1..12.
    """
    months, days = set(), set()
    rows = ws.iter_rows(values_only=True)
    next(rows, None)
    for i, row in enumerate(rows):
        if i >= sample:
            break
        v = row[start_idx]
        if hasattr(v, "month") and v.year > 2000:
            months.add(v.month)
            days.add(v.day)
    return len(days) <= 4 and len(months) >= 6


def read_xlsx_2016(path):
    """2016 ships as an xlsx with the year split across two quarter sheets."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(c) if c is not None else "" for c in next(rows)]
        except StopIteration:
            continue
        cols = {norm(c): i for i, c in enumerate(header)}
        si = next((cols[k] for k in START_KEYS if k in cols), None)
        if si is None:
            print(f"  !! 2016/{sheet}: no start-time column in {header}")
            continue
        ui = next((cols[k] for k in USER_KEYS if k in cols), None)

        swapped = sheet_is_daymonth_swapped(wb[sheet], si)
        buf, total = [], 0
        for row in ws.iter_rows(values_only=True):
            if row[si] is None or row[si] == header[si]:
                continue
            buf.append((row[si], row[ui] if ui is not None else None))
            if len(buf) >= CHUNK:
                total += flush_2016(buf, sheet, swapped)
                buf = []
        if buf:
            total += flush_2016(buf, sheet, swapped)
        print(f"  2016/{sheet.strip()}: {total:,} trips  "
              f"[{'D/M/Y, un-transposed' if swapped else 'as stored'}]"
              f"  (no bike model column)", flush=True)
    wb.close()


def flush_2016(buf, sheet, swapped):
    """Normalise a batch of 2016 cells to ISO strings, then tally them."""
    def to_iso(v):
        if hasattr(v, "strftime"):
            if v.year != 2016:
                return ""          # a stray year-2000 row in the Q3 sheet
            if swapped:            # stored day is the real month, and vice versa
                try:
                    v = v.replace(month=v.day, day=v.month)
                except ValueError:
                    return ""
            return v.strftime("%Y-%m-%d")
        # Text left behind because day > 12 defeated the month-first import.
        ts = pd.to_datetime(str(v), dayfirst=True, errors="coerce")
        return "" if pd.isna(ts) else ts.strftime("%Y-%m-%d")

    df = pd.DataFrame(buf, columns=["start", "user"])
    days = df["start"].map(to_iso)
    keep = days.ne("")
    add(tally, days[keep], classify_user(df.loc[keep, "user"].astype("string")),
        pd.Series(pd.NA, index=df.index[keep], dtype="object"))
    return int(keep.sum())


def walk_zip(path, prefix=""):
    with zipfile.ZipFile(path) as z:
        for info in sorted(z.infolist(), key=lambda i: i.filename):
            name = info.filename
            if name.startswith("__MACOSX") or Path(name).name.startswith("."):
                continue
            low = name.lower()
            if low.endswith(".csv"):
                with z.open(info) as fh:
                    read_csv_stream(fh, prefix + Path(name).name)
            elif low.endswith(".zip"):
                # 2022-11 ships as a zip nested inside the yearly zip.
                walk_zip(io.BytesIO(z.read(info)), prefix + Path(name).stem + "/")


# The City's "Annual Member" label decays out of use and then disappears.
# Checked against the source: the field only ever holds two values, "Annual
# Member" and "Casual Member" -- no third category is being missed. The count
# carrying the first label simply falls away:
#
#     2021-09  annual 287,868 | casual 166,920
#     2021-10  annual 133,942 | casual 233,206   <- flips in one month
#     2023-05  annual  82,577 | casual 506,640
#     2023-08  annual     138 | casual 760,003
#     2023-09  annual       0 | casual 754,303   <- label absent entirely
#     2024-01  schema changes to Member/Casual, and member share returns to 93%
#
# A 63% -> 36% member share in a single month is not a behavioural shift, and
# zero members across four months and 2M trips is impossible. The whole run is
# withheld rather than plotted, from the October 2021 flip to the schema fix.
# `trips` is untouched: only the attribution of trips to rider type is lost.
UNRELIABLE_USER_SPLIT = ("2021-10-01", "2023-12-31")


def suppress_unreliable(df):
    """Blank the member/casual split where the source is provably broken."""
    lo, hi = UNRELIABLE_USER_SPLIT
    span = (df["date"] >= lo) & (df["date"] <= hi)
    df.loc[span, ["trips_member", "trips_casual"]] = ""
    print(f"Withheld the member/casual split for {int(span.sum())} days "
          f"({lo} to {hi}): the source's member label decays to absent.")
    return df


def write_json(df, path):
    """Columnar JSON for the web front-end: same numbers, ~half the bytes of
    row-of-objects, and `cutoff` tells the client where the live API takes over.
    """
    def col(name):
        s = pd.to_numeric(df[name], errors="coerce")
        return [None if pd.isna(v) else int(v) for v in s]

    payload = {
        "cutoff": df["date"].iloc[-1],
        "dates": df["date"].tolist(),
        "trips": col("trips"),
        "member": col("trips_member"),
        "casual": col("trips_casual"),
        "member_classic": col("trips_member_classic"),
        "member_electric": col("trips_member_electric"),
        "casual_classic": col("trips_casual_classic"),
        "casual_electric": col("trips_casual_electric"),
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, separators=(",", ":")))
    print(f"Wrote {path}  ({path.stat().st_size / 1024:.0f} KB)")


def main():
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("bikeshare_daily.csv")
    json_out = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    print("Reducing trip-level archives to daily counts\n")

    xlsx_2016 = DOWNLOADS / "bikeshare-ridership-2016.xlsx"
    if xlsx_2016.exists():
        print(xlsx_2016.name)
        read_xlsx_2016(xlsx_2016)

    for year in range(2017, 2027):
        z = DOWNLOADS / f"bikeshare-ridership-{year}.zip"
        if not z.exists():
            print(f"!! missing {z.name}")
            continue
        print(z.name)
        walk_zip(z)

    rows = []
    for d in sorted(tally):
        c = tally[d]
        has_model = int(d[:4]) in model_years
        row = {
            "date": d,
            "trips": c["trips"],
            "trips_member": c["member"],
            "trips_casual": c["casual"],
        }
        # The joint only exists where a bike model was recorded (2024 on).
        # classic/electric totals are deliberately not stored -- sum the joint.
        for user in ("member", "casual"):
            for model in ("classic", "electric"):
                row[f"trips_{user}_{model}"] = (
                    c[f"{user}_{model}"] if has_model else "")
        rows.append(row)

    df = suppress_unreliable(pd.DataFrame(rows))
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out, index=False)

    print(f"\nWrote {out}  ({len(df):,} days, {out.stat().st_size / 1024:.0f} KB)")
    if json_out:
        write_json(df, json_out)
    print(f"Range: {df['date'].iloc[0]} -> {df['date'].iloc[-1]}")
    print(f"Total trips: {df['trips'].sum():,}")

    dates = pd.to_datetime(df["date"])
    gaps = pd.date_range(dates.min(), dates.max(), freq="D").difference(dates)
    print(f"Days with no data in range: {len(gaps)}"
          + (f" -> {[d.strftime('%Y-%m-%d') for d in gaps[:12]]}" if len(gaps) else ""))

    print("\nPer-year totals:")
    yr = df.assign(y=df["date"].str[:4]).groupby("y").agg(
        days=("date", "size"), trips=("trips", "sum"))
    print(yr.to_string())

    if notes:
        print("\nNotes:")
        for k, v in sorted(notes.items()):
            print(f"  {k}: {v:,}")


if __name__ == "__main__":
    main()
